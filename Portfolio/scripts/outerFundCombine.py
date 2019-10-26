# -*- coding: utf-8 -*-
import os
import sys
import openpyxl
import json

from openpyxl.styles import numbers
from openpyxl.styles import Alignment
# for reduce method of lambda
from functools import reduce
# 统计
from calculateFundCategoryData import calculateFundCategoryData
from model.fundModel import fundModel

class outerFundCombine:
    """
    把天天基金，且慢，螺丝钉计划的数据整合到一张 Excel 表
    """
    def __init__(self):
        self.filenames = [u'danjuan_螺丝钉定投.txt',u'qieman_10万补充ETF计划.txt',u'qieman_我的S定投计划.txt', u'tiantian_康力泉.txt',u'huatai_康力泉.txt',u'guangfa_支付宝.txt']
        self.fundCategorys = self.getFundCategorys()
        
        self.filepaths = []
        # 结果文件路径
        self.resultPath = os.path.join(os.getcwd(), u'output', u'场外基金汇总.xlsx')
        # 先删除旧文件
        if os.path.exists(self.resultPath):
            os.remove(self.resultPath)
        
        for root, dirs, files in os.walk(os.getcwd(), topdown=False):
            for name in files:
                if name in self.filenames:
                    self.filepaths.append(os.path.join(root,name))
        
        # 数据模型集合
        self.fundModelArray = []
        
        #print(self.filepaths)
        outwb = openpyxl.Workbook() # 打开一个将写的文件并创建 sheet 表单
        # 上面的构造函数默认生成一个叫 sheet 的表单，直接 active 属性获取它，并改名就好，不要在单独创建了
        #outws = outwb.create_sheet(title=u'基金持仓')#在将写的文件创建sheet
        outws = outwb.active
        outws.title = u'基金持仓'
        #outwb._active_sheet_index = 1
        # 字体
        font = openpyxl.styles.Font(u'Arial', size = 10, color='000000')
        self.font = font
        # 行游标
        rowCursor = 1
        # 写入标题行
        headerLine = u'基金名称\t基金代码\t持仓成本\t持仓份额\t持仓市值\t累计收益'
        headers = headerLine.split('\t')
        # 加入资产配置分类标签
        headers.append(u'一级分类')
        headers.append(u'二级分类')
        headers.append(u'三级分类')
        headers.append(u'分类 ID')
        # 加入 APP 来源标签
        headers.append(u'来源')
        for i in range(1,len(headers)+1): 
            outws.cell(rowCursor, i).value = headers[i-1]
            outws.cell(rowCursor, i).font = self.font
            align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
            outws.cell(rowCursor, i).alignment = align
        rowCursor = rowCursor + 1
        # 写入基金持仓数据
        for filepath in self.filepaths:
            with open(filepath,'r',encoding='utf-8') as file:
                lines = file.readlines()
                color = self.colorOfFileName(filepath)
                for i in range(2,len(lines)): 
                    values = lines[i].replace('\n','').split('\t')
                    # 加入 model 模型以备统计
                    model = fundModel()
                    #if len(values) == 1:   # 跳过空行
                    #    continue
                    # 基金代码，第二列
                    fundCode = values[1]
                    # 写入基本数据 1 ~ len(values)+1 与资产配置分类数据 +4 与 APP 来源 + 1
                    for col in range(1,len(values)+1 + 4 + 1):
                        outws.cell(rowCursor, col).font = self.font
                        outws.cell(rowCursor, col).fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor=color)
                        if col == 3:    # 持仓净值，保留小数点后 4 位
                            outws.cell(rowCursor,col).number_format = '0.0000'
                            outws.cell(rowCursor, col).value = float(values[col-1])
                        elif col in [4,5,6]:    # 除基金代码，其他数字保留小数点后 2 位
                            outws.cell(rowCursor,col).number_format = '0.00'
                            outws.cell(rowCursor, col).value = float(values[col-1])
                            if col == 5:
                                model.marketCap = float(values[col-1])
                            if col == 6:
                                model.totalGain = float(values[col-1])
                        elif col in [1,2]:
                            outws.cell(rowCursor, col).value = values[col-1]
                        # 分类标签数据
                        elif col in [7,8,9,10]:
                            category = self.getFundCategoryByCode(fundCode)
                            #print(category)
                            if category != '':
                                # col - 6 表示把第七列转化为 category1 即 7 - 6 = 1
                                outws.cell(rowCursor, col).value = category[u'category{0}'.format(col-6)]
                                align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
                                outws.cell(rowCursor, col).alignment = align
                                # model 赋值
                                setattr(model,u'category{0}'.format(col-6),category[u'category{0}'.format(col-6)])
                        # 来源
                        elif col == 11:
                            outws.cell(rowCursor, col).value = self.getAppSourceByFilePath(filepath)
                            align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
                            outws.cell(rowCursor, col).alignment = align
                    self.fundModelArray.append(model)
                    rowCursor = rowCursor + 1
        # 统计环节
        calculator = calculateFundCategoryData()
        # 开始计算并输出
        #calculator.calculate(self.fundModelArray)
        calculator.generateEchartsJson(self.fundModelArray)
        # 保存文件
        outwb.save(self.resultPath)
    
    # 获取资产旭日图分类配置文件
    def getFundCategorys(self):
        path = os.path.join(os.getcwd(),u'config',u'fundCategory.json')
        if not os.path.exists(path):
            print(u'[ERROR] 缺少资产配置分类文件：{0}'.format(path))
            exit()
        
        with open(path,'r',encoding='utf-8') as jsonFile:
            data = json.loads(jsonFile.read())
            fundCategorys = data['data']
            return fundCategorys
            #for category in fundCategorys:
            #    print(category)
    
    # 根据基金代码，获取资产旭日图分类数据
    def getFundCategoryByCode(self,code):
        for fundCategory in self.fundCategorys:
            if code == fundCategory['code']:
                return fundCategory
        return ''
    
    # 根据基金文件，获取 APP 持仓来源
    def getAppSourceByFilePath(self,filepath):
        if u'螺丝钉定投' in filepath:
            return u'螺丝钉定投'
        elif u'10万补充ETF计划' in filepath:
            return u'且慢补充 150 份'
        elif u'我的S定投计划' in filepath:
            return u'且慢 S 定投'       
        elif u'tiantian' in filepath:
            return u'天天基金'
        elif u'guangfa' in filepath:
            return u'支付宝'
        elif u'huatai_康力泉' in filepath:
            return u'股票账户'
        return '未知'
        
    # 不同 APP 配色
    def colorOfFileName(self, name):
        # 色值转换 https://www.sioe.cn/yingyong/yanse-rgb-16/
        if 'danjuan' in name:
            # 242,195,0
            return 'F2C300'
        elif 'qieman' in name:
            # 0,176,204
            return '00B1CC'
        elif 'tiantian' in name:
            # 233,80,26
            return 'E9501A'
        elif 'guangfa' in name:
            # 0,161,233
            return '00A1E9'
        elif 'huatai' in name:
            # 222,48,49
            return 'DE3031'
        else:
            return 'FFFFFF'
            

outerFundCombine = outerFundCombine()