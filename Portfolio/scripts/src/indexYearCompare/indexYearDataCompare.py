# coding=utf-8
import os
import sys
import time
# groupby & itemgetter
from itertools import groupby
from operator import itemgetter

import requests
import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter    # 列宽

# 把父路径加入到 sys.path 供 import 搜索
currentDir = os.path.abspath(os.path.dirname(__file__))
srcDir = os.path.dirname(currentDir)
sys.path.append(srcDir)
from config.pathManager import pathManager
from config.colorConstants import colorConstants

class indexYearModel:

    def __init__(self, data = None):
        self.index = -1 # 序号
        self.name = 'NA'
        self.changeRate = 0.0
        # Excel Properties
        self.fillColor = 'FFFFFF'
        # 支持了一个简易的 json 字符串转 fundModel 对象的逻辑
        if data:
            self.__dict__ = data

    def __str__(self):
        return u'{0}\t{1}\t{2}\t{3}'.format(self.index,self.name,self.changeRate,self.fillColor)

    def __getitem__(self, key):
        return getattr(self, key)
    
    def __setitem__(self, key, value):
        setattr(self,key,value)

"""
获取和对比所有观察品种的年 K 数据，横向对比各年份更好的权益类品种选择
"""
class indexYearDataCompare:

    def __init__(self):
        # 需要比对年线的品种
        self.indexInfos = [ \
            {u"category1":u"A 股",u"category2":u"大盘股",u"category3":u"上证50",u"indexCode":u"000016",u"requestCode":u"0000161",u"categoryId":u"111"}, \
            {u"category1":u"A 股",u"category2":u"大盘股",u"category3":u"50AH",u"indexCode":u"000170",u"requestCode":u"0001701",u"categoryId":u"112"}, \
            {u"category1":u"A 股",u"category2":u"大盘股",u"category3":u"沪深300",u"indexCode":u"000300",u"requestCode":u"0003001",u"categoryId":u"113"}, \
            {u"category1":u"A 股",u"category2":u"大盘股",u"category3":u"300价值",u"indexCode":u"000919",u"requestCode":u"0009191",u"categoryId":u"114"}, \
            {u"category1":u"A 股",u"category2":u"大盘股",u"category3":u"基本面60",u"indexCode":u"399701",u"requestCode":u"3997012",u"categoryId":u"115"}, \
            {u"category1":u"A 股",u"category2":u"大盘股",u"category3":u"基本面120",u"indexCode":u"399702",u"requestCode":u"3997022",u"categoryId":u"116"}, \
            {u"category1":u"A 股",u"category2":u"大盘股",u"category3":u"中小板",u"indexCode":u"399005",u"requestCode":u"3990052",u"categoryId":u"117"}, \
            {u"category1":u"A 股",u"category2":u"中小盘股",u"category3":u"中证500",u"indexCode":u"000905",u"requestCode":u"0009051",u"categoryId":u"121"}, \
            {u"category1":u"A 股",u"category2":u"中小盘股",u"category3":u"中证1000",u"indexCode":u"000852",u"requestCode":u"0008521",u"categoryId":u"123"}, \
            {u"category1":u"A 股",u"category2":u"中小盘股",u"category3":u"创业板",u"indexCode":u"399006",u"requestCode":u"3990062",u"categoryId":u"124"}, \
            {u"category1":u"A 股",u"category2":u"红利价值",u"category3":u"中证红利",u"indexCode":u"000922",u"requestCode":u"0009221",u"categoryId":u"131"}, \
            {u"category1":u"A 股",u"category2":u"行业股",u"category3":u"养老产业",u"indexCode":u"399812",u"requestCode":u"3998122",u"categoryId":u"141"}, \
            {u"category1":u"A 股",u"category2":u"行业股",u"category3":u"全指医药",u"indexCode":u"000991",u"requestCode":u"0009911",u"categoryId":u"142"}, \
            {u"category1":u"A 股",u"category2":u"行业股",u"category3":u"中证环保",u"indexCode":u"000827",u"requestCode":u"0008271",u"categoryId":u"143"}, \
            {u"category1":u"A 股",u"category2":u"行业股",u"category3":u"中证传媒",u"indexCode":u"399971",u"requestCode":u"3999712",u"categoryId":u"144"}, \
            {u"category1":u"A 股",u"category2":u"行业股",u"category3":u"证券公司",u"indexCode":u"399975",u"requestCode":u"3999752",u"categoryId":u"145"}, \
            {u"category1":u"A 股",u"category2":u"行业股",u"category3":u"金融地产",u"indexCode":u"000992",u"requestCode":u"0009921",u"categoryId":u"146"}, \
            {u"category1":u"A 股",u"category2":u"行业股",u"category3":u"全指消费",u"indexCode":u"000990",u"requestCode":u"0009901",u"categoryId":u"147"}, \
            {u"category1":u"海外新兴",u"category2":u"香港",u"category3":u"恒生",u"indexCode":u"HSI5",u"requestCode":u"HSI5",u"categoryId":u"211"}, \
            {u"category1":u"海外新兴",u"category2":u"台湾",u"category3":u"台湾加权",u"indexCode":u"TWII_UI",u"requestCode":u"TWII_UI",u"categoryId":u"213"}, \
            {u"category1":u"海外新兴",u"category2":u"日本",u"category3":u"日经225",u"indexCode":u"N225_UI",u"requestCode":u"N225_UI",u"categoryId":u"214"}, \
            {u"category1":u"海外成熟",u"category2":u"海外成熟",u"category3":u"德国30",u"indexCode":u"GDAXI_UI",u"requestCode":u"GDAXI_UI",u"categoryId":u"311"}, \
            {u"category1":u"海外成熟",u"category2":u"海外成熟",u"category3":u"富时100",u"indexCode":u"FTSE_UI",u"requestCode":u"FTSE_UI",u"categoryId":u"312"}, \
            {u"category1":u"海外成熟",u"category2":u"海外成熟",u"category3":u"法国40",u"indexCode":u"FCHI_UI",u"requestCode":u"FCHI_UI",u"categoryId":u"313"}, \
            {u"category1":u"海外成熟",u"category2":u"海外成熟",u"category3":u"道琼斯",u"indexCode":u"DJIA_UI",u"requestCode":u"DJIA_UI",u"categoryId":u"314"}, \
            {u"category1":u"海外成熟",u"category2":u"海外成熟",u"category3":u"纳斯达克",u"indexCode":u"NDX_UI",u"requestCode":u"NDX_UI",u"categoryId":u"315"}, \
            {u"category1":u"海外成熟",u"category2":u"海外成熟",u"category3":u"标普500",u"indexCode":u"SPX_UI",u"requestCode":u"SPX_UI",u"categoryId":u"316"}, \
            {u"category1":u"债券",u"category2":u"国内债券",u"category3":u"可转债",u"indexCode":u"000832",u"requestCode":u"0008321",u"categoryId":u"411"}, \
            {u"category1":u"商品",u"category2":u"商品",u"category3":u"黄金",u"indexCode":u"GC00Y",u"requestCode":u"GC00Y0",u"categoryId":u"511"}, \
            {u"category1":u"商品",u"category2":u"商品",u"category3":u"原油",u"indexCode":u"CL00Y",u"requestCode":u"CL00Y0",u"categoryId":u"512"}, \
            {u"category1":u"商品",u"category2":u"商品",u"category3":u"白银",u"indexCode":u"SI00Y",u"requestCode":u"SI00Y0",u"categoryId":u"513"} \
        ]

        self.colorConstants = colorConstants()

        self.pm = pathManager()
        self.outputDir = os.path.join(self.pm.outputPath,u'indexYearDataCompare')
        if not os.path.exists(self.outputDir):
            os.makedirs(self.outputDir)

    # 获取所有观察指数的年 K 数据
    def fetchIndexYearData(self):
        for indexInfo in self.indexInfos:
            code = indexInfo['requestCode']
            url = 'http://pdfm2.eastmoney.com/EM_UBG_PDTI_Fast/api/js?id={0}&TYPE=yk'.format(code)
            response = requests.get(url)
            dataList = response.text.split('\r\n')
            indexDir = os.path.join(self.outputDir,u'indexDataOfYear')
            if not os.path.exists(indexDir):
                os.makedirs(indexDir)
            filePath = os.path.join(indexDir,'{0}_{1}_{2}.txt'.format(indexInfo['categoryId'],indexInfo['category3'],indexInfo['indexCode']))

            with open(filePath,'w',encoding='utf-8') as f:
                print(indexInfo['category3'])
                for data in dataList:
                    values = str(data).replace('(','').split(',')
                    if len(values) < 3:
                        continue
                    result = '{0}\t{1}\t{2}\t{3}%'.format(values[0],values[1],values[2],round((float(values[2])/float(values[1])-1)*100,2))
                    print(result)
                    f.write(result + '\n')
                print('\n')
            time.sleep(2)
    
    # 按年度横向比较所有观察指数年 K 的收益表现
    def showIndexYearDataCompare(self, begin=1990):
        # 结束年份 + 1
        endYear = int(time.strftime("%Y", time.localtime())) + 1
        # 年份区间
        years = [x for x in range(begin, endYear)]
        # 文件集合
        filePaths = []
        for root, dirs, files in os.walk(self.outputDir, topdown=False):
            for name in files:
                if '.txt' in name:
                    filePaths.append(os.path.join(root, name))
        compareDir = os.path.join(self.outputDir,u'indexCompareOfYear')
        if not os.path.exists(compareDir):
            os.makedirs(compareDir)
        for year in years:
            print(year)
            with open(os.path.join(compareDir,u'{0}.txt'.format(year)),'w',encoding='utf-8') as outputFile:
                idx = 0
                for filePath in filePaths:
                    with open(filePath,'r',encoding='utf-8') as f:
                        allText = f.read()
                        # 包含对应年份，如 1990-
                        if '{0}-'.format(year) in allText:
                            # 游标回到最开始
                            f.seek(0)
                            lines = f.readlines()
                            name = os.path.basename(filePath).replace('.txt','').split('_')[1]
                            idx = idx + 1
                            for line in lines:
                                if '{0}-'.format(year) in line:
                                    values = line.replace('\n','').split('\t')
                                    print('{0}\t{1}'.format(name,values[3]))
                                    outputFile.write('{0}\t{1}\t{2}'.format(idx, name,values[3]) + '\n')
                                    break
                        else:
                            #print('{0} has no year {1}'.format(filePath,year))
                            continue
            print('\n')
    
    def generateExcelForIndexYearDataCompare(self, begin=1990):
        # 需要先调用 showIndexYearDataCompare 来生成诸如 2005.txt ~ 2019.txt 之类的数据
        compareDir = os.path.join(self.outputDir,u'indexCompareOfYear')
        filePaths = []
        for root, dirs, files in os.walk(compareDir, topdown=False):
            for name in files:
                if '.txt' in name:
                    filePaths.append(os.path.join(root, name))
        # Excel
        outwb = openpyxl.Workbook()
        outws = outwb.active
        # 字体
        font = openpyxl.styles.Font(u'Arial', size = 10, color='333333')

        for filepath in filePaths:
            # 模型数组，用来决定颜色输出
            indexYearModels = []
            #print(filepath)
            # worksheet 命名
            name = os.path.basename(filepath)
            year = int(u'{0}'.format(name.replace('.txt','')))
            if year < begin:
                continue
            #print(year)
            # 文件转 models
            with open(filepath,'r',encoding='utf-8') as f:
                for line in f.readlines():
                    values = line.replace('\n','').replace('%','').split('\t')
                    indexModel = indexYearModel()
                    indexModel.index = int(values[0])
                    indexModel.name = values[1]
                    indexModel.changeRate = round(float(values[2])/100,4)
                    indexYearModels.append(indexModel)
            #数据排序（从高到低）
            indexYearModels.sort(key=itemgetter('changeRate'),reverse=True)
            raiseModels = [x for x in indexYearModels if x.changeRate >= 0]
            failModels = [x for x in indexYearModels if x.changeRate < 0]
            raiseCount = len(raiseModels)
            failCount = len(failModels)
            # 填写上涨品种的颜色值
            if raiseCount > 0:
                step = round((float(100) / raiseCount) / 100,4)
                #print('上涨步进',step)
                current = 1.0
                # 上涨颜色
                for i in range(0,raiseCount):
                    indexYear = indexYearModels[i]
                    indexYear.fillColor = '{0}'.format(self.colorConstants.getGradationColorForRise(current))
                    current = current - step
                    #print(indexYear)
            if failCount > 0:
                step = round((float(100) / failCount) / 100,4)
                #print('下跌步进',step)
                current = 0.0
                # 上涨颜色
                for i in range(raiseCount, len(indexYearModels)):
                    indexYear = indexYearModels[i]
                    indexYear.fillColor = '{0}'.format(self.colorConstants.getGradationColorForFail(current))
                    current = current + step
                    #print(indexYear)
            # 恢复索引排序
            indexYearModels.sort(key=itemgetter('index'))
            outws.title = u'{0}'.format(name.replace('.txt',''))
            # 行游标
            rowCursor = 1
            # 标题
            headers = '序号\t名称\t年涨跌幅'.split('\t')
            for col in range(1,len(headers)+1):
                outws.cell(rowCursor, col).font = font
                outws.cell(rowCursor, col).value = headers[col-1]
            rowCursor = rowCursor + 1            
            # 写入数据
            for model in indexYearModels:
                for col in range(1,4):
                    outws.cell(rowCursor, col).font = font
                    if col == 1:
                        outws.cell(rowCursor, col).value = model.index
                        outws.cell(rowCursor, col).number_format = '0'
                    if col == 2:
                        outws.cell(rowCursor, col).value = model.name
                    if col == 3:
                        outws.cell(rowCursor, col).value = float(model.changeRate)
                        outws.cell(rowCursor, col).number_format = '0.00%'
                        outws.cell(rowCursor, col).fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor=model.fillColor)
                rowCursor = rowCursor + 1
            # 新 worksheet
            outws = outwb.create_sheet(u'new sheet')
            outwb._active_sheet_index = outwb._active_sheet_index + 1
        # 选中第一个
        outwb._active_sheet_index = 0
        # 删除最后一个多于的 sheet
        del outwb[u'new sheet']
        outwb.save(os.path.join(compareDir,'indexYearDataCompare.xlsx'))

if __name__ == "__main__":
    obj = indexYearDataCompare()
    #obj.fetchIndexYearData()
    begin = 1990
    obj.showIndexYearDataCompare(begin=begin)
    obj.generateExcelForIndexYearDataCompare(begin=begin)