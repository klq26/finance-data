# -*- coding: utf-8 -*-
import os
import json

import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter    # 列宽

# model
from model.assetModel import assetModel
# config
from config.assetCategoryConstants import assetCategoryConstants
from config.pathManager import pathManager

class assetAllocationExcelParser:

    def __init__(self):
        self.pm = pathManager()
        self.fundCategorys = self.getFundCategorys()
        categoryConstants = assetCategoryConstants()
        self.category1Array = categoryConstants.category1Array
        self.category2Array = categoryConstants.category2Array
        self.category3Array = categoryConstants.category3Array
        self.modelArray = []

    # 格式化浮点数
    def beautify(self,num):
        return round(float(num),2)

    # 不同 APP 配色
    def getFundColorByAppSourceName(self, name):
        # 色值转换 https://www.sioe.cn/yingyong/yanse-rgb-16/
        if name in [u'螺丝钉定投',u'李淑云螺丝钉',u'康世海螺丝钉']:
            # 242,195,0
            return 'F2C300'
        elif name in [u'且慢补充 150 份',u'且慢 S 定投']:
            # 0,176,204
            return '00B1CC'
        elif u'天天基金' in name:
            # 233,80,26
            return 'E9501A'
        elif u'支付宝' in name:
            # 0,161,233
            return '00A1E9'
        elif u'股票账户' in name:
            # 222,48,49
            return 'DE3031'
        elif u'现金账户' in name:
            # 0,161,233
            return 'F7A128'
        elif u'冻结资金' in name:
            # 222,48,49
            return '8B8C90'
        else:
            return 'FFFFFF'

    # 获取资产旭日图分类配置文件
    def getFundCategorys(self):
        path = os.path.join(self.pm.configPath,u'fundCategory.json')
        if not os.path.exists(path):
            print(u'[ERROR] 缺少资产配置分类文件：{0}'.format(path))
            exit()
        with open(path,'r',encoding='utf-8') as jsonFile:
            data = json.loads(jsonFile.read())
            fundCategorys = data['data']
            return fundCategorys

    # 根据基金代码，获取资产旭日图分类数据
    def getFundCategoryByCode(self,code):
        for fundCategory in self.fundCategorys:
            if code == fundCategory['code']:
                return fundCategory
        return ''

    # 读取 txt，生成 assetModel 基金数据集合，输出到 xlsx 文件
    def generateExcelFile(self,assetModelArray,path=''):
        # 先删除旧文件
        if os.path.exists(path):
            os.remove(path)
    
        outwb = openpyxl.Workbook() # 打开一个将写的文件并创建 sheet 表单
        # 上面的构造函数默认生成一个叫 sheet 的表单，直接 active 属性获取它，并改名就好，不要在单独创建了
        #outws = outwb.create_sheet(title=u'基金持仓')#在将写的文件创建sheet
        #outwb._active_sheet_index = 1
        outws = outwb.active
        outws.title = u'资产配置情况'
        # 字体
        font = openpyxl.styles.Font(u'Arial', size = 10, color='000000')
        self.font = font
        # 行游标
        rowCursor = 1
        # 写入标题行
        headerLine = u'基金名称\t基金代码\t持仓成本\t持仓份额\t持仓市值\t累计收益'
        headers = headerLine.split('\t')
        # 加入资产配置分类标签
        headers.append(u'一级分类')
        headers.append(u'二级分类')
        headers.append(u'三级分类')
        headers.append(u'分类 ID')
        # 加入 APP 来源标签
        headers.append(u'来源')
        for i in range(1,len(headers)+1): 
            outws.cell(rowCursor, i).value = headers[i-1]
            outws.cell(rowCursor, i).font = self.font
            align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
            outws.cell(rowCursor, i).alignment = align
        rowCursor = rowCursor + 1
        # 写入基金持仓数据
        for assetModel in assetModelArray:
            #print(assetModel.__dict__)
            color = self.getFundColorByAppSourceName(assetModel.appSource)
            for col in range(1,len(headers)+1):
                outws.cell(rowCursor, col).font = self.font
                outws.cell(rowCursor, col).fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor=color)
                if col == 1:
                    outws.cell(rowCursor, col).value = assetModel.fundName
                if col == 2:
                    outws.cell(rowCursor,col).number_format = '000000'
                    outws.cell(rowCursor, col).value = int(assetModel.fundCode)
                if col == 3:
                    outws.cell(rowCursor,col).number_format = '0.0000'
                    outws.cell(rowCursor, col).value = float(assetModel.holdNetValue)
                if col == 4:
                    outws.cell(rowCursor,col).number_format = '0.00'
                    outws.cell(rowCursor, col).value = float(assetModel.holdShareCount)
                if col == 5:
                    outws.cell(rowCursor,col).number_format = '0.00'
                    outws.cell(rowCursor, col).value = float(assetModel.holdMarketCap)
                if col == 6:
                    outws.cell(rowCursor,col).number_format = '0.00'
                    outws.cell(rowCursor, col).value = float(assetModel.holdTotalGain)
                # 分类标签数据
                if col in [7,8,9,10]:
                    category = self.getFundCategoryByCode(assetModel.fundCode)
                    if category != '':
                        # col - 6 表示把第七列转化为 category1 即 7 - 6 = 1
                        outws.cell(rowCursor, col).value = category[u'category{0}'.format(col-6)]
                        align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
                        outws.cell(rowCursor, col).alignment = align
                # 来源
                if col == 11:
                    outws.cell(rowCursor, col).value = assetModel.appSource
                    align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
                    outws.cell(rowCursor, col).alignment = align
            rowCursor = rowCursor + 1
        # 自动列宽
        # 获取每一列的内容的最大宽度
        i = 0
        col_width = []
        # 每列
        for column in outws.columns:
            # 每行
            for row in range(len(column)):
                if row == 0:
                    # 数组增加第一个元素 
                    # 注意：这里应该将 str 字符串 decode 成 bytes，因为汉字双字节，ascii 单字节，转成 bytes 才知道谁最长
                    # 所以下面有 encode('utf-8') 就是 unicode -> bytes 的操作
                    col_width.append(len(str(column[row].value).encode('utf-8')))
                else:
                    # 获得每列中的内容的最大宽度
                    if col_width[i] < len(str(column[row].value).encode('utf-8')):
                        col_width[i] = len(str(column[row].value).encode('utf-8'))
            i = i + 1
        #设置列宽
        for i in range(len(col_width)):
            # 根据列的数字返回字母
            col_letter = get_column_letter(i+1)
            # 当宽度大于100，宽度设置为100
            if col_width[i] > 100:
                outws.column_dimensions[col_letter].width = 100
            # 只有当宽度大于10，才设置列宽
            elif col_width[i] > 10:
                outws.column_dimensions[col_letter].width = col_width[i] * 0.75
        # 保存文件
        outwb.save(path)

if __name__ == "__main__":
    assetExcel = assetAllocationExcelParser()
    # 读取文件
    assetModelArray = list
    assetJsonPath = os.path.join(assetExcel.pm.holdingOutputPath, u'{0}asset.json'.format(u'康力泉整体'))
    with open(assetJsonPath,'r',encoding=u'utf-8') as assetJsonFile:
        # object_hook 配合 init 传入 self.__dict__ = dictData 实现 json 字符串转 python 自定义对象
        assetModelArray = json.loads(assetJsonFile.read(),object_hook=assetModel)
    
    assetExcel.generateExcelFile(assetModelArray,path=os.path.join(assetExcel.pm.holdingOutputPath, u'{0}资产配置.xlsx'.format(u'康力泉整体')))