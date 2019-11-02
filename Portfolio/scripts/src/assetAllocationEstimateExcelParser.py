# -*- coding: utf-8 -*-

import os
import sys
import time
import json

import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles import Alignment

from config.pathManager import pathManager
from estimateFundManager import estimateFundManager
from model.fundModel import fundModel
# pretty table output
from prettytable import PrettyTable

class assetAllocationEstimateExcelParser:

    def __init__(self, strategy='a'):
        if strategy == 'a':
            self.pm = pathManager(strategyName='康力泉')
        elif strategy == 'b':
            self.pm = pathManager(strategyName='父母')
        self.fundCategorys = self.getFundCategorys()
        self.fundJsonFilePathExt = ''
        
        # 获取资产旭日图分类配置文件
    def getFundCategorys(self):
        path = os.path.join(self.pm.configPath,u'fundCategory.json')
        if not os.path.exists(path):
            print(u'[ERROR] 缺少资产配置分类文件：{0}'.format(path))
            exit()
        with open(path,'r',encoding='utf-8') as jsonFile:
            data = json.loads(jsonFile.read())
            fundCategorys = data['data']
            return fundCategorys
    
    # 根据基金代码，获取资产旭日图分类数据
    def getFundCategoryByCode(self,code):
        for fundCategory in self.fundCategorys:
            if code == fundCategory['code']:
                return fundCategory
        return ''
    
    def getFundColorByAppSourceName(self, name):
        # 色值转换 https://www.sioe.cn/yingyong/yanse-rgb-16/
        if name in [u'螺丝钉定投',u'李淑云螺丝钉',u'康世海螺丝钉']:
            # 242,195,0
            return 'F2C300'
        elif name in [u'且慢补充 150 份',u'且慢 S 定投']:
            # 0,176,204
            return '00B1CC'
        elif u'天天基金' in name:
            # 233,80,26
            return 'E9501A'
        elif u'支付宝' in name:
            # 0,161,233
            return '00A1E9'
        elif u'股票账户' in name:
            # 222,48,49
            return 'DE3031'
        elif u'现金账户' in name:
            # 0,161,233
            return 'F7A128'
        elif u'冻结资金' in name:
            # 222,48,49
            return '8B8C90'
        else:
            return 'FFFFFF'

    def generateEstimateExcelFile(self, fundModelArray, path=''):
        if len(fundModelArray) == 0:
            return
        # 天天基金获取估算净值
        manager = estimateFundManager()
        headerLine = u'基金名称\t基金代码\t持仓成本\t持仓份额\t持仓市值\t累计收益\t当前净值\t净值日期\t估算净值\t估算涨跌幅\t估算日收益\t估算时间\t一级分类\t二级分类\t三级分类\t分类 ID\t来源'
        # 先删除旧文件
        if os.path.exists(path):
            os.remove(path)

        outwb = openpyxl.Workbook()
        outws = outwb.active
        outws.title = u'基金估算情况'
        # 字体
        font = openpyxl.styles.Font(u'Arial', size = 10, color='000000')
        self.font = font
        # 行游标
        rowCursor = 1
        # 写入标题行
        headers = headerLine.split('\t')
        for i in range(1,len(headers)+1): 
            outws.cell(rowCursor, i).value = headers[i-1]
            outws.cell(rowCursor, i).font = self.font
            align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
            outws.cell(rowCursor, i).alignment = align
        rowCursor = rowCursor + 1
        # 统计
        currentTotalMarketCap = 0.0             # 当前资产整体市值
        currentTotalStockMarketCap = 0.0        # 当前股票市值
        estimateTotalGainToday = 0.0            # 今日涨跌额估值
        outerFundEstimateTotalGainToday = 0.0   # 今日场外涨跌额估值
        gainByAppSource = {}                    # 分 APP 涨跌幅统计
        # esitmate 暂存区，如果一个代码查过了，就不要二次出现浪费时间了
        estimateCache = {}
        # 写入基金持仓数据
        for fundModel in fundModelArray:
            # 即便是现金，也计入总体市值，为了看投资组合的整体收益情况
            currentTotalMarketCap = currentTotalMarketCap + fundModel.holdMarketCap
            if fundModel.category1 in [u'现金',u'冻结资金']:
                continue
            # 只计算权益类资产时使用
            currentTotalStockMarketCap = currentTotalStockMarketCap + fundModel.holdMarketCap
            #print(fundModel.__dict__)
            color = self.getFundColorByAppSourceName(fundModel.appSource)
            if fundModel.fundCode in estimateCache.keys():
                estimateValues = estimateCache[fundModel.fundCode]
            else:
                estimateValues = manager.estimate(fundModel.fundCode)
                estimateCache[fundModel.fundCode] = estimateValues
            #print(estimateValues)  # 元组数据
            time.sleep(0.25)
            if estimateValues:
                fundModel.currentNetValue = estimateValues[0]
                fundModel.currentNetValueDate = estimateValues[1]
                fundModel.estimateNetValue = estimateValues[2]
                fundModel.estimateRate = estimateValues[3]
                fundModel.estimateTime = estimateValues[4]
            else:
                print('{0}估值请求失败'.format(fund.fundCode))
            # 写数据
            for col in range(1,len(headers)+1):
                outws.cell(rowCursor, col).font = self.font
                outws.cell(rowCursor, col).fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor=color)
                if col == 1:
                    outws.cell(rowCursor, col).value = fundModel.fundName
                if col == 2:
                    outws.cell(rowCursor,col).number_format = '000000'
                    outws.cell(rowCursor, col).value = int(fundModel.fundCode)
                if col == 3:
                    outws.cell(rowCursor,col).number_format = '0.0000'
                    outws.cell(rowCursor, col).value = float(fundModel.holdNetValue)
                if col == 4:
                    outws.cell(rowCursor,col).number_format = '0.00'
                    outws.cell(rowCursor, col).value = float(fundModel.holdShareCount)
                if col == 5:
                    outws.cell(rowCursor,col).number_format = '0.00'
                    outws.cell(rowCursor, col).value = float(fundModel.holdMarketCap)
                if col == 6:
                    outws.cell(rowCursor,col).number_format = '0.00'
                    outws.cell(rowCursor, col).value = float(fundModel.holdTotalGain)
                if col == 7:
                    outws.cell(rowCursor,col).number_format = '0.0000'
                    outws.cell(rowCursor, col).value = float(fundModel.currentNetValue)
                if col == 8:
                    outws.cell(rowCursor, col).value = fundModel.currentNetValueDate
                    outws.cell(rowCursor, col).alignment = Alignment(horizontal='center')
                if col == 9:
                    outws.cell(rowCursor,col).number_format = '0.0000'
                    outws.cell(rowCursor, col).value = float(fundModel.estimateNetValue)
                if col == 10:
                    outws.cell(rowCursor,col).number_format = '0.00%'
                    outws.cell(rowCursor, col).value = round(fundModel.estimateRate,4)
                if col == 11:
                    outws.cell(rowCursor,col).number_format = '0.00'
                    # 红涨绿跌
                    changeValue = round((fundModel.estimateNetValue - fundModel.currentNetValue)*fundModel.holdShareCount,2)
                    font = openpyxl.styles.Font(u'Arial', size = 10,bold = True, color='FFFFFF')
                    # http://www.yuangongju.com/color
                    changeValueColor = 'DD2200'
                    if changeValue >= 0:
                        # 221,34,0
                        changeValueColor = 'DD2200'
                    else:
                        # 0,153,51
                        changeValueColor = '009933'
                    outws.cell(rowCursor, col).fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor=changeValueColor)
                    outws.cell(rowCursor, col).font = font
                    outws.cell(rowCursor, col).value = changeValue
                    # 计入今日统计
                    if fundModel.appSource != u'股票账户':
                        # 仅计入场外基金部分（因为场内有情绪涨跌溢价，收盘价不能用净值估算代表）
                        outerFundEstimateTotalGainToday = outerFundEstimateTotalGainToday + changeValue
                    estimateTotalGainToday = estimateTotalGainToday + changeValue
                    # 按 APP 来源统计收支
                    if fundModel.appSource not in gainByAppSource.keys():
                        gainByAppSource[fundModel.appSource] = round(changeValue,2)
                    else:
                        gainOfCurrentAppSource = gainByAppSource[fundModel.appSource]
                        gainByAppSource[fundModel.appSource] = round(gainOfCurrentAppSource + changeValue,2)
                    
                if col == 12:
                    outws.cell(rowCursor, col).value = fundModel.estimateTime
                    outws.cell(rowCursor, col).alignment = Alignment(horizontal='center')
                if col in [13,14,15,16]:
                    category = self.getFundCategoryByCode(fundModel.fundCode)
                    if category != '':
                        # col - 6 表示把第七列转化为 category1 即 13 - 12 = 1
                        outws.cell(rowCursor, col).value = category[u'category{0}'.format(col-12)]
                        align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
                        outws.cell(rowCursor, col).alignment = align
                if col == 17:
                    outws.cell(rowCursor, col).value = fundModel.appSource
                    align = Alignment(horizontal='center') # ,vertical='center',wrap_text=True
                    outws.cell(rowCursor, col).alignment = align
            rowCursor = rowCursor + 1
        # 保存文件
        outwb.save(path)
        # 输出统计
        print('\n场外基金持仓估值结果：\n')
        #print(u'今日预估收益：{0}\t今日场外预估收益：{1}\t今日场内预估收益：{2}'.format(round(estimateTotalGainToday,2), round(outerFundEstimateTotalGainToday,2),round(estimateTotalGainToday - outerFundEstimateTotalGainToday,2)))
        #print(u'之前组合总市值：{0}\t组合预估涨跌幅：{1}%'.format(round(currentTotalMarketCap,2),round(estimateTotalGainToday/currentTotalMarketCap * 100, 2)))
        #print(u'之前权益类总市值：{0}\t权益类预估涨跌幅：{1}%'.format(round(currentTotalStockMarketCap,2),round(estimateTotalGainToday/currentTotalStockMarketCap * 100, 2)))
        # 输出估算整体情况
        tb1 = PrettyTable()
        tb1.field_names = [u'今日收益估算',u'场内估算',u'场外估算',u'权益类涨跌',u'组合涨跌']
        tb1.add_row([round(estimateTotalGainToday,2),round(estimateTotalGainToday - outerFundEstimateTotalGainToday,2),round(outerFundEstimateTotalGainToday,2),u'{0}%'.format(round(estimateTotalGainToday/currentTotalStockMarketCap * 100, 2)),u'{0}%'.format(round(estimateTotalGainToday/currentTotalMarketCap * 100, 2))])
        print(tb1)
        print('\n')
        # 输出分账户情况
        tb2 = PrettyTable()
        tb2.field_names = gainByAppSource.keys()
        tb2.add_row(gainByAppSource.values())
        print(tb2)

    # 读取本地 fundModel 数据
    def loadFundModelArrayFromJson(self):
        # 读取文件
        fundJsonPath = os.path.join(self.pm.holdingOutputPath, u'{0}fund.json'.format(self.fundJsonFilePathExt))
        with open(fundJsonPath,'r',encoding=u'utf-8') as fundJsonFile:
            # object_hook 配合 init 传入 self.__dict__ = dictData 实现 json 字符串转 python 自定义对象
            contentList = json.loads(fundJsonFile.read(),object_hook=fundModel)
            return contentList

if __name__ == '__main__':
    strategy = 'a'
    if len(sys.argv) >= 2:
        #print(u'[ERROR] 参数不足。需要键入策略编号。a：康力泉 b：父母')
        strategy = sys.argv[1]
    if strategy == 'a':
        estimateExcel = assetAllocationEstimateExcelParser('a')
        estimateExcel.fundJsonFilePathExt = u'康力泉整体'
    elif strategy == 'b':
        estimateExcel = assetAllocationEstimateExcelParser('b')
        estimateExcel.fundJsonFilePathExt = u'父母'
        path=os.path.join(estimateExcel.pm.holdingOutputPath, u'{0}收益估算.xlsx'.format(estimateExcel.fundJsonFilePathExt))
        estimateExcel.generateEstimateExcelFile(estimateExcel.loadFundModelArrayFromJson(),path)
    